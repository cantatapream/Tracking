"""
SPT 데이터 매니저 - JSON 저장/로드, Excel/CSV Export, 작전 관리
"""
import json
import os
import time
import datetime
import shutil
from typing import List, Optional
from core.models import Personnel, Equipment, DEFAULT_VESSELS, DEFAULT_PERSONNEL


DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
OPERATIONS_DIR = os.path.join(DATA_DIR, "operations")
STATUS_FILE = os.path.join(DATA_DIR, "status.json")

RANK_ORDER = {"총경": 0, "경정": 1, "경감": 2, "경위": 3, "경사": 4, "경장": 5, "순경": 6}


class DataManager:
    def __init__(self):
        self.personnel: List[Personnel] = []
        self.equipment: List[Equipment] = []
        self.vessels: dict = dict(DEFAULT_VESSELS)
        self.vessel_order: list = []
        self.logs: List[dict] = []
        self.rescue_records: list = []
        self.operation_title: str = ""
        self.ui_settings: dict = {}
        self.custom_dept_name: str = "기타"
        self._created_at: float = 0.0
        self._current_file: str = STATUS_FILE
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(OPERATIONS_DIR, exist_ok=True)

    # ---- 작전 관리 ----
    def new_operation(self):
        """새 작전 시작 - 직전 작전의 인원/장비/단정 유지, 중국어선 제외, 모든 위치 본함"""
        ops = self.list_operations()

        if ops:
            # 직전 작전에서 인원/장비/단정 가져오기
            last_file = ops[0]["filepath"]
            self._load_from_file(last_file)

            # 인원: 모두 본함으로, 타이머/이동내역 초기화
            for p in self.personnel:
                p.location = "base"
                p.status = "standby"
                p.deploy_timestamp = None
                p.last_return_timestamp = None
                p.total_deploy_seconds = 0.0
                p.has_been_deployed = False
                p.movement_history = []

            # 장비: 모두 본함으로, 타이머 초기화
            for e in self.equipment:
                e.vessel_id = ""
                e.is_running = False
                e.run_start_timestamp = None
                e.total_run_seconds = 0.0

            # 중국어선 제거, 단정과 본함만 유지
            self.vessels = {vid: vinfo for vid, vinfo in self.vessels.items()
                           if vinfo["type"] != "vessel"}

            # 로그 및 구조기록 초기화
            self.logs = []
            self.rescue_records = []
        else:
            self._init_defaults()

        self.operation_title = ""
        self._created_at = time.time()
        ts = time.strftime("%Y%m%d_%H%M%S")
        self._current_file = os.path.join(OPERATIONS_DIR, f"op_{ts}.json")
        self.save()

    def load_operation(self, filepath: str) -> bool:
        """과거 작전 불러오기"""
        self._current_file = filepath
        return self._load_from_file(filepath)

    def list_operations(self) -> List[dict]:
        """과거 작전 목록 (최근 순)"""
        ops = []
        for f in os.listdir(OPERATIONS_DIR):
            if f.endswith(".json"):
                filepath = os.path.join(OPERATIONS_DIR, f)
                try:
                    with open(filepath, "r", encoding="utf-8") as fh:
                        data = json.load(fh)
                    ops.append({
                        "title": data.get("operation_title", "제목 없음"),
                        "filepath": filepath,
                        "created_at": data.get("created_at", data.get("last_saved", 0)),
                        "last_saved": data.get("last_saved", 0),
                        "personnel_count": len(data.get("personnel", [])),
                        "log_count": len(data.get("logs", [])),
                    })
                except (json.JSONDecodeError, OSError):
                    pass
        ops.sort(key=lambda x: x["last_saved"], reverse=True)
        return ops

    def delete_operation(self, filepath: str):
        if os.path.exists(filepath) and filepath != self._current_file:
            os.remove(filepath)

    def set_title(self, title: str):
        self.operation_title = title
        self.save()

    # ---- 데이터 로드/저장 ----
    def load(self) -> bool:
        if not os.path.exists(STATUS_FILE):
            self._init_defaults()
            return False
        return self._load_from_file(STATUS_FILE)

    def _load_from_file(self, filepath: str) -> bool:
        if not os.path.exists(filepath):
            self._init_defaults()
            return False
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            self.personnel = [Personnel.from_dict(p) for p in data.get("personnel", [])]
            self.equipment = [Equipment.from_dict(e) for e in data.get("equipment", [])]
            self.vessels = data.get("vessels", dict(DEFAULT_VESSELS))
            self.vessel_order = data.get("vessel_order", [])
            self.logs = data.get("logs", [])
            self.rescue_records = data.get("rescue_records", [])
            self.operation_title = data.get("operation_title", "")
            self._created_at = data.get("created_at", data.get("last_saved", 0))

            for log in self.logs:
                if "type" not in log:
                    log["type"] = "auto"

            old_memos = data.get("memos", [])
            for memo in old_memos:
                self.logs.append({
                    "timestamp": memo["timestamp"],
                    "time_str": memo["time_str"],
                    "message": memo.get("text", ""),
                    "type": "memo"
                })
            if old_memos:
                self.logs.sort(key=lambda x: x.get("timestamp", 0))

            self.ui_settings = data.get("ui_settings", {})
            self.custom_dept_name = data.get("custom_dept_name", "기타")
            self._current_file = filepath
            return True
        except (json.JSONDecodeError, KeyError):
            self._init_defaults()
            return False

    def _init_defaults(self):
        self.personnel = [Personnel(
            id=p.id, name=p.name, rank=p.rank, department=p.department
        ) for p in DEFAULT_PERSONNEL]
        self.equipment = []
        self.vessels = dict(DEFAULT_VESSELS)
        self.logs = []
        self.rescue_records = []
        self.operation_title = ""
        self.ui_settings = {}
        self.custom_dept_name = "기타"

    def save(self):
        data = {
            "personnel": [p.to_dict() for p in self.personnel],
            "equipment": [e.to_dict() for e in self.equipment],
            "vessels": self.vessels,
            "vessel_order": getattr(self, 'vessel_order', []),
            "logs": self.logs,
            "rescue_records": self.rescue_records,
            "operation_title": self.operation_title,
            "created_at": self._created_at or time.time(),
            "last_saved": time.time(),
            "ui_settings": getattr(self, 'ui_settings', {}),
            "custom_dept_name": getattr(self, 'custom_dept_name', '기타'),
        }
        with open(self._current_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    # ---- 로그 ----
    def add_log(self, message: str, log_type: str = "auto"):
        ts = time.time()
        current_date = time.strftime("%y.%m.%d", time.localtime(ts))

        # 날짜 변경 시 자동 구분선 삽입
        if self.logs:
            last_non_sep = None
            for log in reversed(self.logs):
                if log.get("type") != "date_separator":
                    last_non_sep = log
                    break
            if last_non_sep:
                last_date = time.strftime("%y.%m.%d", time.localtime(last_non_sep.get("timestamp", 0)))
                if last_date != current_date:
                    self.logs.append({
                        "timestamp": ts - 0.001,
                        "time_str": "",
                        "message": f"<{current_date}>",
                        "type": "date_separator",
                    })

        entry = {
            "timestamp": ts,
            "time_str": time.strftime("%H:%M:%S"),
            "message": message,
            "type": log_type,
        }
        self.logs.append(entry)
        self.save()
        return entry

    def add_memo(self, text: str):
        return self.add_log(text, log_type="memo")

    def edit_log(self, log_entry: dict, new_message: str):
        log_entry["message"] = new_message
        self.save()

    def delete_log(self, log_entry: dict):
        if log_entry in self.logs:
            idx = self.logs.index(log_entry)
            log_entry["_deleted_index"] = idx
            self.logs.remove(log_entry)
            self.save()

    def restore_log(self, log_entry: dict):
        """삭제된 로그 복원 (원래 위치에 삽입)"""
        idx = log_entry.pop("_deleted_index", len(self.logs))
        idx = min(idx, len(self.logs))
        self.logs.insert(idx, log_entry)
        self.save()

    # ---- 인원 ----
    def get_personnel_by_id(self, pid: str) -> Optional[Personnel]:
        for p in self.personnel:
            if p.id == pid:
                return p
        return None

    def get_personnel_at(self, location: str) -> List[Personnel]:
        result = [p for p in self.personnel if p.location == location]
        return sorted(result, key=lambda p: RANK_ORDER.get(p.rank, 99))

    def get_equipment_at(self, location: str) -> List[Equipment]:
        if location == "base":
            return [e for e in self.equipment if e.vessel_id in ("", "base")]
        return [e for e in self.equipment if e.vessel_id == location]

    def move_equipment_batch(self, eids: List[str], target_location: str) -> Optional[str]:
        names = []
        for eid in eids:
            eq = next((e for e in self.equipment if e.id == eid), None)
            if eq:
                old_loc = eq.vessel_id or "base"
                if old_loc != target_location:
                    eq.vessel_id = target_location
                    names.append(eq.name)
        if not names:
            return None
        target_name = self.get_location_display_name(target_location)
        if target_location == "base":
            target_name = "본함"
        log_msg = f"{target_name}으로 {', '.join(names)} 이동 조치"
        self.add_log(log_msg)
        self.save()
        return log_msg

    def update_personnel_dept(self, pid: str, department: str):
        person = self.get_personnel_by_id(pid)
        if person:
            person.department = department
            self.save()

    def add_personnel(self, name: str, rank: str, department: str = "항해") -> Personnel:
        max_num = 0
        for p in self.personnel:
            try:
                num = int(p.id.replace("P", ""))
                max_num = max(max_num, num)
            except ValueError:
                pass
        new_id = f"P{max_num + 1:03d}"
        person = Personnel(id=new_id, name=name, rank=rank, department=department)
        self.personnel.append(person)
        self.save()
        return person

    def update_personnel(self, pid: str, name: str, rank: str):
        person = self.get_personnel_by_id(pid)
        if person:
            person.name = name
            person.rank = rank
            self.save()

    def remove_personnel(self, pid: str):
        self.personnel = [p for p in self.personnel if p.id != pid]
        self.save()

    def _record_movement(self, person: Personnel, target_location: str):
        """이동 내역 기록"""
        person.movement_history.append({
            "timestamp": time.time(),
            "to": target_location,
            "to_name": self.get_location_display_name(target_location),
        })

    def move_personnel(self, pid: str, target_location: str) -> Optional[str]:
        person = self.get_personnel_by_id(pid)
        if not person:
            return None
        old_loc = person.deploy_to(target_location)
        self._record_movement(person, target_location)
        old_name = self.get_location_display_name(old_loc)
        new_name = self.get_location_display_name(target_location)
        log_msg = f"{person.name} {person.rank} : {old_name} → {new_name}"
        self.add_log(log_msg)
        self.save()
        return log_msg

    def move_personnel_batch(self, pids: List[str], target_location: str) -> Optional[str]:
        names = []
        for pid in pids:
            person = self.get_personnel_by_id(pid)
            if person and person.location != target_location:
                person.deploy_to(target_location)
                self._record_movement(person, target_location)
                names.append(f"{person.rank} {person.name}")
        if not names:
            return None
        target_name = self.get_location_display_name(target_location)
        target_info = self.vessels.get(target_location, {})
        target_type = target_info.get("type", "")
        count_str = f" 총 {len(names)}명"
        name_list = ", ".join(names)
        if target_type == "patrol":
            # 단정 이동
            log_msg = f"{name_list}{count_str} {target_name}으로 이동"
        elif target_type == "vessel":
            # 선박 등선
            log_msg = f"{name_list}{count_str} {target_name} 등선 완료"
        elif target_location == "base":
            log_msg = f"{name_list}{count_str} 본함 복귀"
        else:
            log_msg = f"{name_list}{count_str} → {target_name}"
        self.add_log(log_msg)
        self.save()
        return log_msg

    def get_location_display_name(self, location: str) -> str:
        if location in self.vessels:
            return self.vessels[location]["name"]
        return location

    # ---- 장비 ----
    def add_equipment(self, name: str, category: str = "", vessel_id: str = "") -> Equipment:
        max_num = 0
        for e in self.equipment:
            try:
                num = int(e.id.replace("E", ""))
                max_num = max(max_num, num)
            except ValueError:
                pass
        new_id = f"E{max_num + 1:03d}"
        eq = Equipment(id=new_id, name=name, category=category, vessel_id=vessel_id)
        self.equipment.append(eq)
        self.save()
        return eq

    def remove_equipment(self, eid: str):
        self.equipment = [e for e in self.equipment if e.id != eid]
        self.save()

    def assign_equipment(self, eid: str, assignee_id: Optional[str]):
        for e in self.equipment:
            if e.id == eid:
                old = e.assignee_id
                e.assignee_id = assignee_id
                if assignee_id:
                    person = self.get_personnel_by_id(assignee_id)
                    pname = person.name if person else assignee_id
                    self.add_log(f"장비 '{e.name}' 담당자 지정: {pname}")
                else:
                    self.add_log(f"장비 '{e.name}' 담당자 해제")
                self.save()
                return old
        return None

    # ---- 선박 ----
    def add_vessel(self, vessel_id: str, name: str, vessel_type: str):
        self.vessels[vessel_id] = {"name": name, "type": vessel_type}
        # 어선인 경우 순서 목록에 추가
        if vessel_type == "vessel":
            if not hasattr(self, 'vessel_order') or self.vessel_order is None:
                self.vessel_order = []
            if vessel_id not in self.vessel_order:
                self.vessel_order.append(vessel_id)
        self.save()

    def get_vessel_order(self) -> list:
        """어선 표시 순서 반환"""
        if not hasattr(self, 'vessel_order') or not self.vessel_order:
            # 기존 어선을 이름순으로 초기화
            self.vessel_order = sorted(
                [vid for vid, v in self.vessels.items() if v["type"] == "vessel"]
            )
        # 삭제된 vessel 제거, 새로 추가된 vessel 추가
        existing = {vid for vid, v in self.vessels.items() if v["type"] == "vessel"}
        self.vessel_order = [vid for vid in self.vessel_order if vid in existing]
        for vid in existing:
            if vid not in self.vessel_order:
                self.vessel_order.append(vid)
        return self.vessel_order

    def set_vessel_order(self, order: list):
        """어선 표시 순서 저장"""
        self.vessel_order = order
        self.save()

    def remove_vessel(self, vessel_id: str):
        if vessel_id in self.vessels and vessel_id != "base":
            for p in self.get_personnel_at(vessel_id):
                self.move_personnel(p.id, "base")
            # 해당 선박의 장비도 본함으로
            for e in self.equipment:
                if e.vessel_id == vessel_id:
                    e.vessel_id = ""
            del self.vessels[vessel_id]
            if hasattr(self, 'vessel_order') and vessel_id in self.vessel_order:
                self.vessel_order.remove(vessel_id)
            self.save()

    # ---- 구조 기록 ----
    def add_rescue_record(self, data: dict) -> dict:
        """구조/인계/인수 기록 추가"""
        # ID 생성
        max_num = 0
        for r in self.rescue_records:
            try:
                num = int(r["id"].replace("R", ""))
                max_num = max(max_num, num)
            except (ValueError, KeyError):
                pass
        new_id = f"R{max_num + 1:03d}"
        record = {
            "id": new_id,
            "type": data.get("type", "rescue"),
            "timestamp": data.get("timestamp", ""),
            "location": data.get("location", ""),
            "name": data.get("name", ""),
            "gender": data.get("gender", "남"),
            "age": data.get("age", "미상"),
            "severity": data.get("severity", "지연"),
            "initial_state": data.get("initial_state", ""),
            "treatment": data.get("treatment", ""),
            "transferred": data.get("transferred", False),
            "transfer_target": data.get("transfer_target", ""),
            "transfer_timestamp": data.get("transfer_timestamp", ""),
            "source_record_id": data.get("source_record_id", ""),
        }
        self.rescue_records.append(record)
        self.save()
        return record

    def get_rescue_records(self, filter_type=None) -> list:
        """구조 기록 조회 (타입 필터링)"""
        if filter_type is None:
            return list(self.rescue_records)
        return [r for r in self.rescue_records if r.get("type") == filter_type]

    def update_rescue_record(self, record_id: str, field: str, value):
        """구조 기록 특정 필드 업데이트"""
        for r in self.rescue_records:
            if r.get("id") == record_id:
                r[field] = value
                self.save()
                return r
        return None

    def delete_rescue_record(self, record_id: str):
        """구조 기록 삭제"""
        self.rescue_records = [r for r in self.rescue_records if r.get("id") != record_id]
        self.save()

    def get_next_unknown_name(self) -> str:
        """다음 미상 이름 반환 (미상1, 미상2, ...)"""
        max_num = 0
        for r in self.rescue_records:
            name = r.get("name", "")
            if name.startswith("미상"):
                try:
                    num = int(name.replace("미상", ""))
                    max_num = max(max_num, num)
                except ValueError:
                    pass
        return f"미상{max_num + 1}"

    # ---- 내보내기 ----
    def _filter_logs_by_date(self, filter_date=None):
        """일자별 로그 필터링"""
        if not filter_date:
            return self.logs
        return [
            log for log in self.logs
            if datetime.date.fromtimestamp(log.get("timestamp", 0)) == filter_date
        ]

    def export_csv(self, filepath: str, filter_date=None):
        import csv
        logs = self._filter_logs_by_date(filter_date)
        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["일자", "시각", "유형", "내용"])
            for log in logs:
                log_date = datetime.date.fromtimestamp(log.get("timestamp", 0)).strftime("%Y.%m.%d")
                log_type = "메모" if log.get("type") == "memo" else "작전"
                writer.writerow([log_date, log["time_str"], log_type, log["message"]])

    def export_xlsx(self, filepath: str, filter_date=None):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            self.export_csv(filepath.replace(".xlsx", ".csv"), filter_date=filter_date)
            return

        logs = self._filter_logs_by_date(filter_date)

        wb = Workbook()
        ws_log = wb.active
        ws_log.title = "작전 로그"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1B2838", end_color="1B2838", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = ["번호", "일자", "시각", "유형", "내용"]
        for col, h in enumerate(headers, 1):
            cell = ws_log.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for i, log in enumerate(logs, 1):
            log_date = datetime.date.fromtimestamp(log.get("timestamp", 0)).strftime("%Y.%m.%d")
            log_type = "메모" if log.get("type") == "memo" else "작전"
            ws_log.cell(row=i + 1, column=1, value=i).border = thin_border
            ws_log.cell(row=i + 1, column=2, value=log_date).border = thin_border
            ws_log.cell(row=i + 1, column=3, value=log["time_str"]).border = thin_border
            ws_log.cell(row=i + 1, column=4, value=log_type).border = thin_border
            ws_log.cell(row=i + 1, column=5, value=log["message"]).border = thin_border

        ws_log.column_dimensions["A"].width = 8
        ws_log.column_dimensions["B"].width = 12
        ws_log.column_dimensions["C"].width = 12
        ws_log.column_dimensions["D"].width = 8
        ws_log.column_dimensions["E"].width = 50

        ws_per = wb.create_sheet("인원 현황")
        per_headers = ["ID", "이름", "계급", "현재 위치", "상태", "누적 이함(분)"]
        for col, h in enumerate(per_headers, 1):
            cell = ws_per.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for i, p in enumerate(self.personnel, 1):
            loc_name = self.get_location_display_name(p.location)
            ws_per.cell(row=i + 1, column=1, value=p.id).border = thin_border
            ws_per.cell(row=i + 1, column=2, value=p.name).border = thin_border
            ws_per.cell(row=i + 1, column=3, value=p.rank).border = thin_border
            ws_per.cell(row=i + 1, column=4, value=loc_name).border = thin_border
            ws_per.cell(row=i + 1, column=5, value=p.status).border = thin_border
            ws_per.cell(row=i + 1, column=6, value=round(p.total_deploy_seconds / 60, 1)).border = thin_border

        wb.save(filepath)
